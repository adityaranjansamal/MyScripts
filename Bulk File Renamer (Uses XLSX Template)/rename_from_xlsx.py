#!/usr/bin/env python3

import os
import sys
from pathlib import Path
from openpyxl import load_workbook

if len(sys.argv) != 3:
    print("Usage:")
    print("python3 rename_from_xlsx.py <folder> <rename_map.xlsx>")
    sys.exit(1)

folder = Path(sys.argv[1]).expanduser().resolve()
xlsx = Path(sys.argv[2]).expanduser().resolve()

wb = load_workbook(xlsx, data_only=True)
ws = wb.active

renamed = 0
skipped = 0
errors = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row:
        continue

    old_name = str(row[0]).strip() if row[0] else ""
    new_name = str(row[1]).strip() if row[1] else ""

    if not old_name or not new_name:
        continue

    old_file = folder / old_name
    new_file = folder / new_name

    if not old_file.exists():
        print(f"[SKIP] Missing: {old_name}")
        skipped += 1
        continue

    if new_file.exists():
        print(f"[SKIP] Already exists: {new_name}")
        skipped += 1
        continue

    try:
        old_file.rename(new_file)
        print(f"[OK] {old_name}")
        renamed += 1
    except Exception as e:
        print(f"[ERROR] {old_name}")
        print(e)
        errors += 1

print("\n==========")
print(f"Renamed : {renamed}")
print(f"Skipped : {skipped}")
print(f"Errors  : {errors}")
